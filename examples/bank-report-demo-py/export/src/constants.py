from pydantic import BaseModel, Field, create_model
from typing import List, Dict, Union, Any, Optional
import os

# Agent lib and functions dependencies
import requests
import openpyxl

# Agent libs
async def private_fn():
  return True


# Agent functions
def get_bank_branches(bank_id):
    url = f'https://banks.data.fdic.gov/api/locations?filters=CERT%3A{bank_id}&fields=NAME%2CUNINUM%2CSERVTYPE%2CRUNDATE%2CCITY%2CSTNAME%2CZIP%2CCOUNTY%2CADDRESS%2CMAINOFF&sort_by=NAME&sort_order=DESC&limit=10000&offset=0&format=json&download=false'
    response = requests.get(url)
    response_data = response.json()
    return [
        {
            'id': branch['data']['ID'],
            'address': branch['data']['ADDRESS'],
            'city': branch['data']['CITY'],
            'county': branch['data']['COUNTY'],
            'state': branch['data']['STNAME'],
            'zip': branch['data']['ZIP'],
        } for branch in response_data['data']
    ]


def get_bank_financials(bank_id):
    url = f'https://banks.data.fdic.gov/api/financials?filters=CERT%3A{bank_id}&fields=CERT%2CREPDTE%2CASSET%2CDEP&sort_by=REPDTE&sort_order=DESC&limit=10&offset=0&agg_by=REPDTE&agg_sum_fields=DEP&agg_limit=1000&format=json&download=false&filename=data_file'
    response = requests.get(url)
    response_data = response.json()
    return [
        {
            'report_date': f"{r['data']['REPDTE'][:4]}-{r['data']['REPDTE'][4:6]}-{r['data']['REPDTE'][6:]}",
            'total_deposits': r['data']['sum_DEP'] * 1000
        } for r in response_data['data']
    ]


def get_bank_id_by_name(name):
    url = f'https://banks.data.fdic.gov/api/institutions?filters=ACTIVE%3A1&search=NAME:{requests.utils.quote(name)}&fields=NAME'
    response = requests.get(url)
    response_data = response.json()
    return response_data['data'][0]['data']['ID']

def create_workbook():
    return openpyxl.Workbook()


def auto_size_column_width(worksheet):
    for column in worksheet.columns:
        lengths = [len(str(cell.value)) for cell in column if cell.value]
        max_length = max(lengths, default=10)
        column_letter = column[0].column_letter
        worksheet.column_dimensions[column_letter].width = max_length


def add_worksheet(workbook, sheet_name):
    return workbook.create_sheet(title=sheet_name)


def add_table_rows(worksheet, start_row, start_col, rows, formats=None):
    for row_index, row_data in enumerate(rows):
        current_row = start_row + row_index
        for cell_index, cell_data in enumerate(row_data):
            cell = worksheet.cell(row=current_row, column=start_col + cell_index)
            cell.value = cell_data
            if formats and formats[cell_index]:
                cell.number_format = formats[cell_index]


def add_table_header(worksheet, row, col, column_names):
    for index, column_name in enumerate(column_names):
        cell = worksheet.cell(row=row, column=col + index)
        cell.value = column_name
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
        cell.border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'), 
            left=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin'), 
            right=openpyxl.styles.Side(style='thin')
        )


class TASKS:
    @staticmethod
    def bank_excel_report(bank_name: str):
        import json
        import openpyxl
        import os
        
        bank_id = get_bank_id_by_name(bank_name)
        financials = get_bank_financials(bank_id)
        
        workbook = create_workbook()
        worksheet = add_worksheet(workbook, "Financial Report")
        
        headers = ["Report Date", "Total Deposits"]
        rows = [[record["report_date"], record["total_deposits"]] for record in financials]
        
        add_table_header(worksheet, 1, 1, headers)
        add_table_rows(worksheet, 2, 1, rows)
        auto_size_column_width(worksheet)
        
        file_name = f"{bank_name.replace(' ', '_')}_Financial_Report.xlsx"
        file_path = os.getcwd() + '/' + file_name
        workbook.save(file_path)
        
        result = {
            "result": "Financial report generated successfully.",
            "files": [
                {
                    "path": file_path,
                    "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                }
            ]
        }
        
        print(json.dumps(result))
    
    @staticmethod
    def bank_report_by_year(bank_name: str, years: str):
        import json
    
        bank_id = get_bank_id_by_name(bank_name)
        financials = get_bank_financials(bank_id)
        
        selected_years = set(int(year) for year in years.split('|'))
        
        report = [
            {"Report Date": record["report_date"], "Total Deposits": record["total_deposits"]}
            for record in financials if int(record["report_date"][:4]) in selected_years
        ]
        
        print(json.dumps(report))
    
    @staticmethod
    def bank_simple_report(bank_name: str):
        import json
        
        bank_id = get_bank_id_by_name(bank_name)
        financials = get_bank_financials(bank_id)
        
        report = [
            {"Report Date": record["report_date"], "Total Deposits": record["total_deposits"]}
            for record in financials
        ]
        
        print(json.dumps(report))
    

TASK_TOOL_SCHEMAS = {
  "bank_excel_report": {
  "description": "Generates a financial report in Excel format for a specified bank.",
  "returns_description": "Object<{ result: string; files: List<{ path: string; mimeType: string }> }>",
  "input": {"bank_name": str},
  "args_schema": create_model(
    "BankExcelReport",
    bank_name=(str, ...)  # This defines 'bank_name' as a required field of type str
  ),
  "output": Dict[str, Any],
  "function": TASKS.bank_excel_report
},
"bank_report_by_year": {
  "description": "Generates a financial report for a specific bank given a list of specific years",
  "returns_description": "List<{ 'Report Date': string; 'Total Deposits': integer; }>",
  "input": {"bank_name": str, "years": str},
  "args_schema": create_model(
    "BankReportByYear",
    bank_name=(str, ...),
    years=(str, ...)
  ),
  "output": List[Dict[str, Union[str, int]]],
  "function": TASKS.bank_report_by_year
},
"bank_simple_report": {
  "description": "Generates a financial report for a specific bank",
  "returns_description": "Array<{ Report Date: string; Total Deposits: int; }>",
  "input": {"bank_name": str},
  "args_schema": create_model(
    "bankSimpleReport",
    bank_name=(str, ...) 
  ),
  "output": List[Dict[str, Union[str, int]]],
  "function": TASKS.bank_simple_report
}
}

COMPLETION_PROMPT_TEXT = """You are a helpful technical assistant

# ASSISTANT INSTRUCTIONS
This is an agent specialized in retrieving bank data from the FDIC API

# GUIDELINES FOR USING TOOLS AND GENERATING RESPONSES

- Apply your best judgment to decide which tasks to run, if one or more tasks look like they do the same pick a single one
- You must use tool run_adhoc_task at least once if you can't get the results you need with other tools
- To answer questions give preference to tasks that don't generate files unless the user specifically asks for them
- If the task response includes file paths append them to the end of your response as described in the json block instructions below
- For math formulas use syntax supported by KaTex and use $$ as delimiter
- Escape any $s that appear in output so they aren't interpreted as katex markdown
- If the user doesn't explicitly ask for a file, asume the data should be rendered with markdown in the response itself
- Always use markdown to format your response, prefer tables and text formatting over code blocks unless its code
- Be strict about the accuracy of your responses, always use the data you get from tools to answer the user's question
- If you cannot answer the question solely from the tools results, reply with a friendly error message explaining that you don't have the necessary information or capabilities to answer the question
- Use all of the tool results data unless specifically told to subset, summarize or use only part of the data
- Avoid making assumptions or providing speculative answers, when in doubt ask for clarification

# CRITERIA FOR USING TOOLS

- If none of the existing tools help you fulfill the request, use the run_adhoc_task tool to fulfill the request
- When using run_adhoc_task, make your best guess to select the most suitable agent based on its description and tools
- If the run_adhoc_task result doesn't fully address the user's request or seems incorrect, try using run_adhoc_task again with a refined task description (more details below)
- Only after exhausting all possibilities with run_adhoc_task, if you still cannot provide accurate information, reply with a friendly error message explaining that you don't have the necessary information or capabilities to answer the question

# AD-HOC TASK INSTRUCTIONS
- Try your best to use existing tools but if there aren't any that can be used to fulfill the user's request then call run_adhoc_task to achieve what you need to do, select the most suitable agent based on its description and existing tools
- Look suspiciously at results that are not what you expect: run_adhoc_task generates and runs new code and the results could be wrong, apply your best judgment to determine if the result looks correct or not
    - For example: it returned an array with only invalid or missing data like nulls or empty strings
- If the results do not look correct try to fix them by using run_adhoc_task again with an updated description of the task
- When possible prefer using run_adhoc_task with a description that will get you markdown formatted results over raw data and return to the user as-is
"""

TASK_NAME_TO_FUNCTION_NAME_MAP = {
  "bank-excel-report": "bank_excel_report",
  "bank-report-by-year": "bank_report_by_year",
  "bank-simple-report": "bank_simple_report"
}

ADHOC_PROMPT_TEXT = """
You have these globally available public functions:

```
- g

- e

- t

- _

- b

- a

- n

- k

- _

- b

- r

- a

- n

- c

- h

- e

- s

- (

- b

- a

- n

- k

- _

- i

- d

- :

-  

- i

- n

- t

- )

-  

- :

-  

- L

- i

- s

- t

- [

- D

- i

- c

- t

- [

- s

- t

- r

- ,

-  

- U

- n

- i

- o

- n

- [

- s

- t

- r

- ,

-  

- i

- n

- t

- ]

- ]

- ]

-  

- -

-  

- F

- e

- t

- c

- h

- e

- s

-  

- a

-  

- l

- i

- s

- t

-  

- o

- f

-  

- b

- a

- n

- k

-  

- b

- r

- a

- n

- c

- h

- e

- s

-  

- f

- o

- r

-  

- a

-  

- g

- i

- v

- e

- n

-  

- b

- a

- n

- k

- _

- i

- d

-  

- f

- r

- o

- m

-  

- t

- h

- e

-  

- F

- D

- I

- C

-  

- A

- P

- I

- .

-  

- R

- e

- t

- u

- r

- n

- s

-  

- a

-  

- l

- i

- s

- t

-  

- o

- f

-  

- d

- i

- c

- t

- i

- o

- n

- a

- r

- i

- e

- s

- ,

-  

- e

- a

- c

- h

-  

- c

- o

- n

- t

- a

- i

- n

- i

- n

- g

-  

- '

- i

- d

- '

- ,

-  

- '

- a

- d

- d

- r

- e

- s

- s

- '

- ,

-  

- '

- c

- i

- t

- y

- '

- ,

-  

- '

- c

- o

- u

- n

- t

- y

- '

- ,

-  

- '

- s

- t

- a

- t

- e

- '

- ,

-  

- a

- n

- d

-  

- '

- z

- i

- p

- '

-  

- o

- f

-  

- b

- r

- a

- n

- c

- h

- e

- s

- .

- 


- 


- g

- e

- t

- _

- b

- a

- n

- k

- _

- f

- i

- n

- a

- n

- c

- i

- a

- l

- s

- (

- b

- a

- n

- k

- _

- i

- d

- :

-  

- i

- n

- t

- )

-  

- :

-  

- L

- i

- s

- t

- [

- D

- i

- c

- t

- [

- s

- t

- r

- ,

-  

- U

- n

- i

- o

- n

- [

- s

- t

- r

- ,

-  

- i

- n

- t

- ]

- ]

- ]

-  

- -

-  

- F

- e

- t

- c

- h

- e

- s

-  

- b

- a

- n

- k

-  

- f

- i

- n

- a

- n

- c

- i

- a

- l

- s

-  

- f

- o

- r

-  

- a

-  

- g

- i

- v

- e

- n

-  

- b

- a

- n

- k

- _

- i

- d

-  

- f

- r

- o

- m

-  

- t

- h

- e

-  

- F

- D

- I

- C

-  

- A

- P

- I

- .

-  

- R

- e

- t

- u

- r

- n

- s

-  

- a

-  

- l

- i

- s

- t

-  

- o

- f

-  

- d

- i

- c

- t

- i

- o

- n

- a

- r

- i

- e

- s

- ,

-  

- e

- a

- c

- h

-  

- c

- o

- n

- t

- a

- i

- n

- i

- n

- g

-  

- '

- r

- e

- p

- o

- r

- t

- _

- d

- a

- t

- e

- '

-  

- a

- n

- d

-  

- '

- t

- o

- t

- a

- l

- _

- d

- e

- p

- o

- s

- i

- t

- s

- '

- .

- 


- 


- g

- e

- t

- _

- b

- a

- n

- k

- _

- i

- d

- _

- b

- y

- _

- n

- a

- m

- e

- (

- n

- a

- m

- e

- :

-  

- s

- t

- r

- )

-  

- :

-  

- i

- n

- t

-  

- -

-  

- F

- e

- t

- c

- h

- e

- s

-  

- t

- h

- e

-  

- b

- a

- n

- k

- _

- i

- d

-  

- f

- o

- r

-  

- a

-  

- g

- i

- v

- e

- n

-  

- b

- a

- n

- k

-  

- n

- a

- m

- e

-  

- f

- r

- o

- m

-  

- t

- h

- e

-  

- F

- D

- I

- C

-  

- A

- P

- I

- .

-  

- R

- e

- t

- u

- r

- n

- s

-  

- t

- h

- e

-  

- I

- D

-  

- o

- f

-  

- t

- h

- e

-  

- b

- a

- n

- k

- .

- 


- 


- c

- r

- e

- a

- t

- e

- _

- w

- o

- r

- k

- b

- o

- o

- k

- (

- )

-  

- :

-  

- o

- p

- e

- n

- p

- y

- x

- l

- .

- W

- o

- r

- k

- b

- o

- o

- k

-  

- -

-  

- C

- r

- e

- a

- t

- e

- s

-  

- a

- n

- d

-  

- r

- e

- t

- u

- r

- n

- s

-  

- a

-  

- n

- e

- w

-  

- O

- p

- e

- n

- p

- y

- x

- l

-  

- W

- o

- r

- k

- b

- o

- o

- k

-  

- i

- n

- s

- t

- a

- n

- c

- e

- .

- 


- 


- a

- u

- t

- o

- _

- s

- i

- z

- e

- _

- c

- o

- l

- u

- m

- n

- _

- w

- i

- d

- t

- h

- (

- w

- o

- r

- k

- s

- h

- e

- e

- t

- :

-  

- o

- p

- e

- n

- p

- y

- x

- l

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- W

- o

- r

- k

- s

- h

- e

- e

- t

- )

-  

- :

-  

- N

- o

- n

- e

-  

- -

-  

- A

- d

- j

- u

- s

- t

- s

-  

- t

- h

- e

-  

- c

- o

- l

- u

- m

- n

-  

- w

- i

- d

- t

- h

-  

- o

- f

-  

- a

- n

-  

- O

- p

- e

- n

- p

- y

- x

- l

-  

- w

- o

- r

- k

- s

- h

- e

- e

- t

-  

- b

- a

- s

- e

- d

-  

- o

- n

-  

- t

- h

- e

-  

- l

- e

- n

- g

- t

- h

-  

- o

- f

-  

- t

- h

- e

-  

- c

- o

- n

- t

- e

- n

- t

-  

- i

- n

-  

- e

- a

- c

- h

-  

- c

- o

- l

- u

- m

- n

- .

- 


- 


- a

- d

- d

- _

- w

- o

- r

- k

- s

- h

- e

- e

- t

- (

- w

- o

- r

- k

- b

- o

- o

- k

- :

-  

- o

- p

- e

- n

- p

- y

- x

- l

- .

- W

- o

- r

- k

- b

- o

- o

- k

- ,

-  

- s

- h

- e

- e

- t

- _

- n

- a

- m

- e

- :

-  

- s

- t

- r

- )

-  

- :

-  

- o

- p

- e

- n

- p

- y

- x

- l

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- W

- o

- r

- k

- s

- h

- e

- e

- t

-  

- -

-  

- A

- d

- d

- s

-  

- a

-  

- n

- e

- w

-  

- w

- o

- r

- k

- s

- h

- e

- e

- t

-  

- w

- i

- t

- h

-  

- t

- h

- e

-  

- s

- p

- e

- c

- i

- f

- i

- e

- d

-  

- n

- a

- m

- e

-  

- t

- o

-  

- a

- n

-  

- O

- p

- e

- n

- p

- y

- x

- l

-  

- w

- o

- r

- k

- b

- o

- o

- k

- .

-  

- R

- e

- t

- u

- r

- n

- s

-  

- t

- h

- e

-  

- c

- r

- e

- a

- t

- e

- d

-  

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- 


- 


- a

- d

- d

- _

- t

- a

- b

- l

- e

- _

- r

- o

- w

- s

- (

- w

- o

- r

- k

- s

- h

- e

- e

- t

- :

-  

- o

- p

- e

- n

- p

- y

- x

- l

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- W

- o

- r

- k

- s

- h

- e

- e

- t

- ,

-  

- s

- t

- a

- r

- t

- _

- r

- o

- w

- :

-  

- i

- n

- t

- ,

-  

- s

- t

- a

- r

- t

- _

- c

- o

- l

- :

-  

- i

- n

- t

- ,

-  

- r

- o

- w

- s

- :

-  

- L

- i

- s

- t

- [

- L

- i

- s

- t

- [

- U

- n

- i

- o

- n

- [

- s

- t

- r

- ,

-  

- i

- n

- t

- ]

- ]

- ]

- ,

-  

- f

- o

- r

- m

- a

- t

- s

- :

-  

- O

- p

- t

- i

- o

- n

- a

- l

- [

- L

- i

- s

- t

- [

- s

- t

- r

- ]

- ]

-  

- =

-  

- N

- o

- n

- e

- )

-  

- :

-  

- N

- o

- n

- e

-  

- -

-  

- A

- d

- d

- s

-  

- r

- o

- w

- s

-  

- o

- f

-  

- d

- a

- t

- a

-  

- t

- o

-  

- t

- h

- e

-  

- s

- p

- e

- c

- i

- f

- i

- e

- d

-  

- w

- o

- r

- k

- s

- h

- e

- e

- t

-  

- s

- t

- a

- r

- t

- i

- n

- g

-  

- f

- r

- o

- m

-  

- t

- h

- e

-  

- g

- i

- v

- e

- n

-  

- r

- o

- w

-  

- a

- n

- d

-  

- c

- o

- l

- u

- m

- n

- .

-  

- O

- p

- t

- i

- o

- n

- a

- l

- l

- y

-  

- a

- p

- p

- l

- i

- e

- s

-  

- t

- h

- e

-  

- p

- r

- o

- v

- i

- d

- e

- d

-  

- f

- o

- r

- m

- a

- t

- s

-  

- t

- o

-  

- e

- a

- c

- h

-  

- c

- o

- l

- u

- m

- n

- .

- 


- 


- a

- d

- d

- _

- t

- a

- b

- l

- e

- _

- h

- e

- a

- d

- e

- r

- (

- w

- o

- r

- k

- s

- h

- e

- e

- t

- :

-  

- o

- p

- e

- n

- p

- y

- x

- l

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

- W

- o

- r

- k

- s

- h

- e

- e

- t

- ,

-  

- r

- o

- w

- :

-  

- i

- n

- t

- ,

-  

- c

- o

- l

- :

-  

- i

- n

- t

- ,

-  

- c

- o

- l

- u

- m

- n

- _

- n

- a

- m

- e

- s

- :

-  

- L

- i

- s

- t

- [

- s

- t

- r

- ]

- )

-  

- :

-  

- N

- o

- n

- e

-  

- -

-  

- A

- d

- d

- s

-  

- a

-  

- t

- a

- b

- l

- e

-  

- h

- e

- a

- d

- e

- r

-  

- t

- o

-  

- t

- h

- e

-  

- s

- p

- e

- c

- i

- f

- i

- e

- d

-  

- r

- o

- w

-  

- a

- n

- d

-  

- c

- o

- l

- u

- m

- n

-  

- i

- n

-  

- t

- h

- e

-  

- w

- o

- r

- k

- s

- h

- e

- e

- t

- .

-  

- E

- a

- c

- h

-  

- c

- o

- l

- u

- m

- n

-  

- n

- a

- m

- e

-  

- i

- s

-  

- b

- o

- l

- d

- e

- d

- ,

-  

- c

- e

- n

- t

- e

- r

- e

- d

- ,

-  

- a

- n

- d

-  

- b

- o

- r

- d

- e

- r

- e

- d

- .

```

Using only these functions execute the following instructions:

In a codeblock at the top of your response write a python function called doTask that fulfills the given requirements:

- Your only task is to write code for doTask and return the code with no text before or after it.
- You are limited to using the functions described above and operations on the data they return using built-in python functions, otherwise reply with "The request cannot be fulfilled using the available functions" and give a detailed explanation of why.
- When calling the functions only use parameters included in the function definition and be careful to use await only if the function is async.
- Your answer is limited to a single function that only calls the public functions described above, do not use any other functions not included in this set but you don't need to use all of them.
- If you need to import any dependencies for your doTask code always do so inside the doTask function.
- Do not catch errors, let exceptions propagate.
- If there are no errors doTask must always finish by writing its result as JSON to stdout.
- If the output format is not specified, default to JSON to stdout.
- Never output anything else to stdout, any messages if needed should be included in the resulting JSON.
- Do not include any comments or documentation in your code, only the code is needed.
- Remember that you can write code to process the function results to filter or summarize them as needed if the function results are not what is needed.
- If none of the examples given to you are useful for generating the doTask function, generate the code to best of your ability based on the instructions and the available functions.
- Always use print to output information, do not use any other logging mechanism.

- Any values mentioned in the task description should be declared as constants inside the function doTask.
- Be very flexible and proactive, make reasonable assumptions or guesses to extract the parameters needed for the doTask function from the task description and fill in any missing information.
- Do not create files unless explicitly requested, otherwise only output plain text JSON data to stdout.
"""

LIBS = { private_fn, }

FUNCTIONS = { get_bank_branches,
get_bank_financials,
get_bank_id_by_name,
create_workbook,
auto_size_column_width,
add_worksheet,
add_table_rows,
add_table_header, }

IS_LAMBDA = bool(os.getenv('AWS_LAMBDA_FUNCTION_NAME'))

ENV_VARS = {
  "DATA_FILES": "./data"
}
