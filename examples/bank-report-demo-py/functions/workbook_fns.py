import openpyxl

def create_workbook():
    return openpyxl.Workbook()

def auto_size_column_width(worksheet):
    for column in worksheet.columns:
        lengths = [len(str(cell.value)) for cell in column if cell.value]
        max_length = max(lengths, default=10)
        column_letter = column[0].column_letter
        worksheet.column_dimensions[column_letter].width = max_length

def add_worksheet(workbook, sheet_name):
    return workbook.create_sheet(title=sheet_name)

def add_table_rows(worksheet, start_row, start_col, rows, formats=None):
    for row_index, row_data in enumerate(rows):
        current_row = start_row + row_index
        for cell_index, cell_data in enumerate(row_data):
            cell = worksheet.cell(row=current_row, column=start_col + cell_index)
            cell.value = cell_data
            if formats and formats[cell_index]:
                cell.number_format = formats[cell_index]

def add_table_header(worksheet, row, col, column_names):
    for index, column_name in enumerate(column_names):
        cell = worksheet.cell(row=row, column=col + index)
        cell.value = column_name
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
        cell.border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'), 
            left=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin'), 
            right=openpyxl.styles.Side(style='thin')
        )