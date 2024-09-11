##
# DEPENDENCIES
# Warning: these are extracted from your function files, if you need to make changes edit the function file and recompile this task.
##

import requests
import openpyxl
    
##
# LIBRARY FUNCTIONS
# Warning: these are common functions, if you need to make changes edit the function file and recompile this task.
##

async def private_fn():
  return True

##
# PUBLIC FUNCTIONS
# Warning: these are common functions, if you need to make changes edit the function file and recompile this task.
##

def get_bank_branches(bank_id):
    url = f'https://banks.data.fdic.gov/api/locations?filters=CERT%3A{bank_id}&fields=NAME%2CUNINUM%2CSERVTYPE%2CRUNDATE%2CCITY%2CSTNAME%2CZIP%2CCOUNTY%2CADDRESS%2CMAINOFF&sort_by=NAME&sort_order=DESC&limit=10000&offset=0&format=json&download=false'
    response = requests.get(url)
    response_data = response.json()
    return [
        {
            'id': branch['data']['ID'],
            'address': branch['data']['ADDRESS'],
            'city': branch['data']['CITY'],
            'county': branch['data']['COUNTY'],
            'state': branch['data']['STNAME'],
            'zip': branch['data']['ZIP'],
        } for branch in response_data['data']
    ]



def get_bank_financials(bank_id):
    url = f'https://banks.data.fdic.gov/api/financials?filters=CERT%3A{bank_id}&fields=CERT%2CREPDTE%2CASSET%2CDEP&sort_by=REPDTE&sort_order=DESC&limit=10&offset=0&agg_by=REPDTE&agg_sum_fields=DEP&agg_limit=1000&format=json&download=false&filename=data_file'
    response = requests.get(url)
    response_data = response.json()
    return [
        {
            'report_date': f"{r['data']['REPDTE'][:4]}-{r['data']['REPDTE'][4:6]}-{r['data']['REPDTE'][6:]}",
            'total_deposits': r['data']['sum_DEP'] * 1000
        } for r in response_data['data']
    ]



def get_bank_id_by_name(name):
    url = f'https://banks.data.fdic.gov/api/institutions?filters=ACTIVE%3A1&search=NAME:{requests.utils.quote(name)}&fields=NAME'
    response = requests.get(url)
    response_data = response.json()
    return response_data['data'][0]['data']['ID']


def create_workbook():
    return openpyxl.Workbook()



def auto_size_column_width(worksheet):
    for column in worksheet.columns:
        lengths = [len(str(cell.value)) for cell in column if cell.value]
        max_length = max(lengths, default=10)
        column_letter = column[0].column_letter
        worksheet.column_dimensions[column_letter].width = max_length



def add_worksheet(workbook, sheet_name):
    return workbook.create_sheet(title=sheet_name)



def add_table_rows(worksheet, start_row, start_col, rows, formats=None):
    for row_index, row_data in enumerate(rows):
        current_row = start_row + row_index
        for cell_index, cell_data in enumerate(row_data):
            cell = worksheet.cell(row=current_row, column=start_col + cell_index)
            cell.value = cell_data
            if formats and formats[cell_index]:
                cell.number_format = formats[cell_index]



def add_table_header(worksheet, row, col, column_names):
    for index, column_name in enumerate(column_names):
        cell = worksheet.cell(row=row, column=col + index)
        cell.value = column_name
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(vertical='center', horizontal='center')
        cell.border = openpyxl.styles.Border(
            top=openpyxl.styles.Side(style='thin'), 
            left=openpyxl.styles.Side(style='thin'),
            bottom=openpyxl.styles.Side(style='thin'), 
            right=openpyxl.styles.Side(style='thin')
        )

##
# GENERATED CODE
# This function is the generated code: it's safe to edit.
##

def doTask(bank_name: str):
    import json
    import openpyxl
    import os
    
    bank_id = get_bank_id_by_name(bank_name)
    financials = get_bank_financials(bank_id)
    
    workbook = create_workbook()
    worksheet = add_worksheet(workbook, "Financial Report")
    
    headers = ["Report Date", "Total Deposits"]
    rows = [[record["report_date"], record["total_deposits"]] for record in financials]
    
    add_table_header(worksheet, 1, 1, headers)
    add_table_rows(worksheet, 2, 1, rows)
    auto_size_column_width(worksheet)
    
    file_name = f"{bank_name.replace(' ', '_')}_Financial_Report.xlsx"
    file_path = os.getcwd() + '/' + file_name
    workbook.save(file_path)
    
    result = {
        "result": "Financial report generated successfully.",
        "files": [
            {
                "path": file_path,
                "mimeType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        ]
    }
    
    print(json.dumps(result))
